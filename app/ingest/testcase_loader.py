"""テストケース取込・検証（詳細設計§2.3）。

CSV / Excel から targets・cases を読み、構造化して返す。
検証エラーはすべて収集し、TestcaseValidationError でまとめて送出する。
この層は決定的（LLM非依存）。カバレッジ計算・スコアは扱わない（Task 3）。
"""

from __future__ import annotations

import csv
from dataclasses import dataclass

# 許可される技法（targets・cases 両方で検証する）。
ALLOWED_TECHNIQUES = {"境界値分析", "デシジョンテーブル", "状態遷移", "同値分割"}

# 必須列。
TARGET_REQUIRED = ["target_id", "technique", "category", "target"]
CASE_REQUIRED = ["test_id", "technique", "input", "expected", "target_ids"]


@dataclass
class Target:
    target_id: str
    technique: str
    category: str
    target: str
    must_cover: bool = True
    note: str = ""


@dataclass
class Case:
    test_id: str
    technique: str
    precondition: str
    input: str
    expected: str
    target_ids: list[str]
    note: str = ""


class TestcaseValidationError(Exception):
    """検証エラーをまとめて保持する例外。messages 属性に全件を持つ。"""

    # pytest が "Test" 始まりの名前をテストクラスと誤収集しないよう抑止する。
    __test__ = False

    def __init__(self, messages: list[str]):
        self.messages = messages
        super().__init__("テストケース検証エラー:\n" + "\n".join(messages))


def _clean(value) -> str:
    """None を空文字に、前後空白を除去して文字列化する。"""
    if value is None:
        return ""
    return str(value).strip()


def _parse_must_cover(value) -> bool:
    """must_cover は N のときのみ False。空欄や Y は True。"""
    return _clean(value).upper() != "N"


def _parse_target_ids(value) -> list[str]:
    """カンマ区切りを分割し、前後空白除去・空要素を捨てる。"""
    return [t for t in (s.strip() for s in _clean(value).split(",")) if t]


def _build_targets(rows: list[dict]) -> list[Target]:
    return [
        Target(
            target_id=_clean(r.get("target_id")),
            technique=_clean(r.get("technique")),
            category=_clean(r.get("category")),
            target=_clean(r.get("target")),
            must_cover=_parse_must_cover(r.get("must_cover")),
            note=_clean(r.get("note")),
        )
        for r in rows
    ]


def _build_cases(rows: list[dict]) -> list[Case]:
    return [
        Case(
            test_id=_clean(r.get("test_id")),
            technique=_clean(r.get("technique")),
            precondition=_clean(r.get("precondition")),
            input=_clean(r.get("input")),
            expected=_clean(r.get("expected")),
            target_ids=_parse_target_ids(r.get("target_ids")),
            note=_clean(r.get("note")),
        )
        for r in rows
    ]


def _validate(
    target_header: list[str],
    case_header: list[str],
    targets: list[Target],
    cases: list[Case],
) -> list[str]:
    """全エラーを収集して返す。人が直せるよう列名・ID・test_id を含める。"""
    messages: list[str] = []

    # 必須列の不足（ヘッダ単位）。
    for col in TARGET_REQUIRED:
        if col not in target_header:
            messages.append(f"targets: 必須列が不足しています: {col}")
    for col in CASE_REQUIRED:
        if col not in case_header:
            messages.append(f"cases: 必須列が不足しています: {col}")

    # target_id の重複。
    seen: set[str] = set()
    dup: set[str] = set()
    for t in targets:
        if t.target_id in seen:
            dup.add(t.target_id)
        seen.add(t.target_id)
    for d in sorted(dup):
        messages.append(f"targets: target_id が重複しています: {d}")

    # 技法の許可集合チェック（値がある場合のみ）。
    for t in targets:
        if t.technique and t.technique not in ALLOWED_TECHNIQUES:
            messages.append(f"targets: 技法が不正です (target_id={t.target_id}): {t.technique}")
    for c in cases:
        if c.technique and c.technique not in ALLOWED_TECHNIQUES:
            messages.append(f"cases: 技法が不正です (test_id={c.test_id}): {c.technique}")

    # 未定義参照（cases.target_ids が targets に無い ID を指す）。
    valid_ids = {t.target_id for t in targets if t.target_id}
    for c in cases:
        for tid in c.target_ids:
            if tid not in valid_ids:
                messages.append(
                    f"cases: 未定義の target_id を参照しています (test_id={c.test_id}): {tid}"
                )

    return messages


def _assemble(
    target_header: list[str],
    case_header: list[str],
    target_rows: list[dict],
    case_rows: list[dict],
) -> tuple[list[Target], list[Case]]:
    """構造化と検証を行う。CSV・Excel 共通の入口。"""
    targets = _build_targets(target_rows)
    cases = _build_cases(case_rows)
    messages = _validate(target_header, case_header, targets, cases)
    if messages:
        raise TestcaseValidationError(messages)
    return targets, cases


def _read_csv(path: str) -> tuple[list[dict], list[str]]:
    # utf-8-sig で BOM 付き CSV も許容する。
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        header = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]
    return rows, header


def load_from_csv(targets_csv: str, cases_csv: str) -> tuple[list[Target], list[Case]]:
    target_rows, target_header = _read_csv(targets_csv)
    case_rows, case_header = _read_csv(cases_csv)
    return _assemble(target_header, case_header, target_rows, case_rows)


def _read_sheet(wb, name: str) -> tuple[list[dict], list[str]]:
    # シートが無ければ空扱い（必須列不足として検出される）。
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return [], []
    header = [_clean(h) for h in data[0]]
    rows: list[dict] = []
    for r in data[1:]:
        # 全セル空の行はスキップ。
        if all(_clean(v) == "" for v in r):
            continue
        row = {header[i]: r[i] for i in range(len(header)) if i < len(r)}
        rows.append(row)
    return rows, header


def load_from_xlsx(xlsx_path: str) -> tuple[list[Target], list[Case]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        target_rows, target_header = _read_sheet(wb, "targets")
        case_rows, case_header = _read_sheet(wb, "cases")
    finally:
        wb.close()
    return _assemble(target_header, case_header, target_rows, case_rows)


def extract_workbook_text(xlsx_path: str, max_chars: int = 60000) -> str:
    """任意レイアウトのワークブックを、レビュー用に読めるテキストへ抽出する。

    工程ごとにテンプレートが異なる成果物Excel（固定の targets/cases に一致しないもの）を、
    全シート・全行から非空セルを拾って「シートごとの箇条書き」に整形する。
    検証・カバレッジは行わない（決定的なテキスト化のみ）。長すぎる場合は末尾を省略する。
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for name in wb.sheetnames:
            parts.append(f"## シート: {name}")
            for row in wb[name].iter_rows(values_only=True):
                cells = [_clean(c) for c in row]
                # 結合セル由来の空セルが多いので、非空セルだけを区切って1行にする。
                line = " | ".join(c for c in cells if c != "")
                if line:
                    parts.append(line)
            parts.append("")
    finally:
        wb.close()

    text = "\n".join(parts).strip()
    if len(text) > max_chars:
        text = text[:max_chars] + "\n…（以下省略）"
    return text
